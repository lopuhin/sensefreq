#!/usr/bin/env python
import argparse
from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment

from rs.active_dict.loader import get_ad_word


def main():
    parser = argparse.ArgumentParser()
    arg = parser.add_argument
    arg('ad_root')
    arg('contexts_root', type=Path)
    arg('words', type=Path)
    arg('output')
    arg('--limit', type=int, default=100)
    args = parser.parse_args()

    wb = Workbook()
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')

    words = [l.strip() for l in args.words.read_text(encoding='utf8').split('\n')
             if l.strip()]
    for i, word in enumerate(words):
        contexts_path = args.contexts_root / '{}.txt'.format(word)
        if not contexts_path.exists():
            print('Contexts for word "{}" not found, skpping'.format(word))
            continue
        contexts = [l.split('\t') for l in
                    contexts_path.read_text(encoding='utf8').split('\n')]
        contexts = [ctx for ctx in contexts if len(ctx) == 3]
        if not contexts:
            print('No contexts for word "{}", skipping'.format(word))
            continue
        if len(contexts) > args.limit:
            random.seed(42)
            contexts = random.sample(contexts, args.limit)
        else:
            print('Warning: only {} contexts for word "{}"'
                  .format(len(contexts), word))

        ad_word = get_ad_word(word, args.ad_root, with_contexts=False)
        if not ad_word:
            print('Word "{}" not found in AD'.format(ad_word))
            continue

        if i == 0:
            ws = wb.active
            ws.title = word
        else:
            ws = wb.create_sheet(word)

        for row, m in enumerate(ad_word['meanings'], 1):
            ws.cell(row=row, column=3, value='{name}: {meaning}'.format(**m))
            ws.cell(row=row, column=4, value=row)
        n_senses = len(ad_word['meanings'])
        ws.cell(row=n_senses + 1, column=3, value='Другое:')
        ws.cell(row=n_senses + 1, column=4, value=n_senses + 1)
        ws.cell(row=n_senses + 2, column=3, value='Не могу определить:')
        ws.cell(row=n_senses + 2, column=4, value=0)

        for row, (left, center, right) in enumerate(contexts, n_senses + 3):
            ws.cell(row=row, column=1, value=left).alignment = right_align
            ws.cell(row=row, column=2, value=center).alignment = center_align
            ws.cell(row=row, column=3, value=right)
            ws.cell(row=row, column=4, value='-').alignment = right_align
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = \
            2 + max(len(center) for _, center, _ in contexts)
        ws.column_dimensions['C'].width = 80

    wb.save(args.output)


if __name__ == '__main__':
    main()
